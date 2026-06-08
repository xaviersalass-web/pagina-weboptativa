#!/usr/bin/env python3
"""Agrega las respuestas reales de la encuesta a data/survey.json.

Lee "Respuestas encuesta.xlsx" (30 encuestados), calcula la distribución
demográfica y el promedio de cada canción IGNORANDO los 0 (= "no la conozco",
según la guía de Evidencia 2) y exporta un JSON liviano para la web.
"""

import json
import re
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Respuestas encuesta.xlsx"
OUT = ROOT / "data" / "survey.json"

OCUPACION_CORTA = {
    "Trabajador/a independiente / freelance": "Freelance",
    "Hogar/ labores domesticas": "Hogar",
}


def edad_label(v):
    s = str(v).strip()
    if " a " in s:
        return s.replace(" a ", "–")
    m = re.search(r"(\d+)", s)
    return f"{m.group(1)}+" if m else s


def rating(v):
    """Primer dígito de respuestas tipo '4 = Me gusta mucho'."""
    if v is None:
        return None
    m = re.match(r"\s*(\d)", str(v))
    return int(m.group(1)) if m else None


def split_artist_song(header):
    name = re.sub(r"\s*\*+\s*$", "", header).strip()
    parts = re.split(r"\s*[—–-]\s*", name, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return "", name


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True)]
    header = [str(h).strip() if h else "" for h in rows[0]]
    data = [r for r in rows[1:] if any(v is not None for v in r)]

    def col(pred):
        return next((i for i, h in enumerate(header) if pred(h)), None)

    i_edad = col(lambda h: "edad" in h.lower())
    i_genero = col(lambda h: h.strip().lower() == "genero")
    i_ocu = col(lambda h: "ocupaci" in h.lower())
    i_pref = col(lambda h: "géneros" in h.lower() or "preferidos" in h.lower())

    def counter(idx, transform=lambda x: x):
        c = Counter()
        for r in data:
            if r[idx] is not None and str(r[idx]).strip():
                c[transform(str(r[idx]).strip())] += 1
        return c

    edad_c = counter(i_edad, edad_label)
    edad_orden = ["18–24", "25–34", "35–44", "45–54", "55+"]
    edad = [[k, edad_c[k]] for k in edad_orden if edad_c.get(k)]
    edad += [[k, v] for k, v in edad_c.items() if k not in edad_orden]

    genero = counter(i_genero).most_common()
    ocupacion = counter(i_ocu, lambda x: OCUPACION_CORTA.get(x, x)).most_common()

    pref = Counter()
    for r in data:
        if r[i_pref]:
            for g in re.split(r"[;,/]", str(r[i_pref])):
                g = g.strip()
                if g:
                    pref[g] += 1
    generos = pref.most_common()

    # Canciones: columnas con guion entre artista y título
    song_cols = [i for i, h in enumerate(header)
                 if 6 <= i < len(header) - 1 and re.search(r"[—–-]", h)]
    canciones = []
    for i in song_cols:
        artist, song = split_artist_song(header[i])
        vals = [rating(r[i]) for r in data]
        conocen = [v for v in vals if v and v > 0]
        if conocen:
            canciones.append({
                "cancion": song,
                "artista": artist,
                "avg": round(sum(conocen) / len(conocen), 2),
                "conocen": len(conocen),
            })

    top = sorted([c for c in canciones if c["conocen"] >= 5],
                 key=lambda c: c["avg"], reverse=True)[:8]

    survey = {
        "n": len(data),
        "fecha": "mayo 2026",
        "edad": edad,
        "genero": [list(x) for x in genero],
        "ocupacion": [list(x) for x in ocupacion],
        "generos": [list(x) for x in generos],
        "top_canciones": top,
    }
    OUT.write_text(json.dumps(survey, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT.name}: n={survey['n']}, {len(top)} canciones top")
    print(json.dumps(survey, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
